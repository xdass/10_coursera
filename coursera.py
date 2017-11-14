import sys
import requests
import bs4
from openpyxl import Workbook
from lxml import etree

HEADER_FOR_RU = {'accept-language': 'ru-RU,ru;q=0.8,en-US;q=0.6,en;q=0.4'}


def get_courses_url_list(courses_to_parse=20):
    xml_response = requests.get('https://www.coursera.org/sitemap~www~courses.xml')
    tree = etree.fromstring(xml_response.content)
    links = [node[0].text for node in tree]
    return links[:courses_to_parse]


def get_course_page(url):
    html_response = requests.get(url, headers=HEADER_FOR_RU)
    return html_response.content


def parse_course_info(course_html):
    soup = bs4.BeautifulSoup(course_html, 'lxml')
    course_title = soup.select_one('.title').string
    course_start_date = soup.select_one('.startdate span').string
    course_language = soup.select_one('.rc-Language').text
    course_rating = soup.select_one('.ratings-info div:nth-of-type(2)')
    if course_rating:
        course_rating = course_rating.string
    else:
        course_rating = None
    course_weeks = len(soup.select('.rc-WeekView > div'))
    return {
        'course_title': course_title,
        'course_start_date': course_start_date,
        'course_language': course_language,
        'course_rating': course_rating,
        'course_weeks': course_weeks
    }


def collect_courses_info(url_list):
    courses_list = []
    for url in url_list:
        course_page = get_course_page(url)
        courses_list.append(parse_course_info(course_page))
    return courses_list


def fill_excel_workbook(courses):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Coursera courses'
    ws['A1'] = 'Название'
    ws['B1'] = 'Язык'
    ws['C1'] = 'Дата начала'
    ws['D1'] = 'Продолжительность (недели)'
    ws['E1'] = 'Рейтинг'
    offset = 2
    max_row = len(courses) + offset
    rows_count = [number for number in range(offset, max_row)][::-1]
    for course_info in courses:
        row = rows_count.pop()
        ws['A{}'.format(row)] = course_info['course_title']
        ws['B{}'.format(row)] = course_info['course_language']
        ws['C{}'.format(row)] = course_info['course_start_date']
        ws['D{}'.format(row)] = course_info['course_weeks']
        ws['E{}'.format(row)] = course_info['course_rating']
    return wb


def save_to_xlsx(workbook, filepath='coursera_courses.xlsx'):
    workbook.save(filepath)


if __name__ == '__main__':
    print('Parse courses links....')
    courses_links = get_courses_url_list()
    print('Start parsing courses data....')
    courses_info_list = collect_courses_info(courses_links)
    print('Save to file')
    filled_workbook = fill_excel_workbook(courses_info_list)
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
        save_to_xlsx(filled_workbook, filepath)
    else:
        save_to_xlsx(filled_workbook)
